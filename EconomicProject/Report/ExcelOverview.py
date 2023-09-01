import openpyxl
import os
from copy import copy
from Constants import Constants
from openpyxl.styles import Alignment, Font
import time


class ExcelRecorder:
    def __init__(self, file_path):
        self.time = time.time()
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.sheets = self.workbook.sheetnames

        self.verify_rec()
        self.record_sheet = self.workbook.create_sheet("Record")

    def verify_rec(self):
        if "Record" in self.sheets:
            ans = input("Вкладка Record уже существует! Введите 'да' чтобы удалить вкладку, 'нет', чтобы оставить:\n")
            if ans.lower() == "да":
                del self.workbook['Record']
            else:
                raise MemoryError("Вкладка Record уже существует!")

    @staticmethod
    def copy_cell(src_sheet, src_row, src_col,
                  tgt_sheet, tgt_row, tgt_col,
                  copy_style=True):
        cell = src_sheet.cell(src_row, src_col)
        new_cell = tgt_sheet.cell(tgt_row, tgt_col, cell.value)
        if cell.has_style and copy_style:
            new_cell._style = copy(cell._style)

    @staticmethod
    def copy_row(index, ws_src, ws_trg):
        ws_trg_last_row = ws_trg.max_row
        for i, row in enumerate(ws_src.iter_rows(min_row=index, max_row=index), 1):
            for cell in row:
                ExcelRecorder.copy_cell(ws_src, cell.row, cell.column,
                                        ws_trg, ws_trg_last_row + i, cell.column)

    def copy_sheet(self, name):
        ws_trg_last_row = self.record_sheet.max_row
        self.record_sheet.cell(row=ws_trg_last_row, column=1, value=name)

        for index in Constants.COPY_DICT[name]:
            ExcelRecorder.copy_row(index, self.workbook[name], self.record_sheet)

    def copy_sheets(self):
        self.record_sheet.insert_rows(1, 2)
        self.record_sheet.cell(row=1, column=1, value="Длительность проекта")
        self.record_sheet.cell(row=1, column=2, value="20 лет")

        for name in self.sheets:
            self.copy_sheet(name)

    def merge_cells(self):
        for block_name, block in Constants.MERGE_HEAD_DICT.items():
            if block:
                for row in block:
                    self.record_sheet.merge_cells(row)

    def npv_text(self):
        self.record_sheet['A147'] = self.workbook['NPV']['A29'].value
        self.record_sheet['A147'].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        self.record_sheet.merge_cells('A147:G155')

    def copy_graph(self):
        output_dir = "Charts"
        for name, dict_property in Constants.GRAPH_DICT.items():
            for index in dict_property['index']:
                source_chart = self.workbook[name]._charts[index]
                new_chart = copy(source_chart)

                new_chart.anchor = dict_property['place'][index]
                attr_dict = dict_property['attributes'][index]
                for attribute, value in attr_dict.items():
                    setattr(new_chart, attribute, value)

                self.record_sheet.add_chart(new_chart)

    def make_title(self):
        ft = Font(
            name='Calibri',
            size=16,
        )

        align = Alignment(horizontal='center',
                          vertical='center',
                          # text_rotation=0,
                          # wrap_text=False,
                          # shrink_to_fit=False,
                          # indent=0
                          )

        for row in Constants.TITLE_LIST:
            target_cell = row.split(":")[0]
            self.record_sheet[target_cell].font = ft
            self.record_sheet[target_cell].alignment = align
            self.record_sheet.merge_cells(row)

    def set_dimensions(self):
        width_list = ['A', 'B', 'C']
        for column in width_list:
            self.record_sheet.column_dimensions[column].width = 15

        # for size, lst in Constants.height_dict.items():
        #     for row in lst:
        #         self.record_sheet.row_dimensions[row].height = size

    def record(self):
        self.copy_sheets()
        self.merge_cells()
        self.npv_text()
        self.copy_graph()
        self.make_title()
        self.set_dimensions()
        self.__save_changes()

    def __save_changes(self, path="Results/"):
        name = 'Overview_'
        self.workbook.save(path + name + os.path.basename(self.file_path).split('_')[1])

    def __close_workbook(self):
        self.workbook.close()

    def __del__(self):
        print(f"Файл {os.path.basename(self.file_path)} закрыт!")
        print(f"Время обработки составило {round(time.time() - self.time, 2)} c.")
        self.__close_workbook()


if __name__ == "__main__":
    file_path = "../Examples/Report_E2.xlsx"
    wb = ExcelRecorder(file_path)
    wb.record()
