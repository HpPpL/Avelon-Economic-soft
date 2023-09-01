import matplotlib.pyplot as plt
from openpyxl import load_workbook
import time


class ChartsToImage:
    def __init__(self, workbook):
        self.workbook = workbook
        self.time = time.time()

    @staticmethod
    def wrap_row_data(worksheet, index, start_column):
        for row in worksheet.iter_rows(min_row=index, max_row=index, min_col=start_column, values_only=True):
            row_text = "\t".join(str(cell) for cell in row)  # Разделение ячеек табуляцией
            return list(map(lambda x: round(float(x), 2), row_text.split()))

    @staticmethod
    def wrap_column_data(worksheet, column, num=10):
        return [cell[0].value for cell in worksheet.iter_rows(min_col=column, max_col=column)][1:num]

    def npv_to_image(self, path='Charts/'):
        # График сроков окупаемости
        wrap_row_data = ChartsToImage.wrap_row_data
        column = 4
        worksheet = self.workbook["NPV"]
        data = {
            "Накопленный денежный поток": wrap_row_data(worksheet, 7, column),
            "Накопленный дисконтированный денежный поток": wrap_row_data(worksheet, 10, column),
            "Год": wrap_row_data(worksheet, 1, column)
        }

        plt.figure(figsize=(10, 6))
        plt.plot(data["Год"], data["Накопленный денежный поток"], marker='o', color='blue',
                 label="Накопленный денежный поток")
        plt.plot(data["Год"], data["Накопленный дисконтированный денежный поток"], marker='o', color='red',
                 label="Накопленный дисконтированный денежный поток")
        plt.xlabel("Год")
        plt.ylabel("Значение")
        plt.title("График сроков окупаемости")
        plt.legend(loc='upper left', bbox_to_anchor=(1, 1))
        plt.grid(True)

        plt.savefig(path + "график1.png", bbox_inches="tight")

    def elastic_to_image(self, path='Charts/'):
        # График общей чувствительности
        wrap_row_data = ChartsToImage.wrap_row_data
        column = 3
        worksheet = self.workbook["Elastic"]
        data = {
            "CAPEX": wrap_row_data(worksheet, 2, column),
            "OPEX": wrap_row_data(worksheet, 14, column),
            "Market: Продажа на внутреннем рынке": wrap_row_data(worksheet, 25, column),
            "Market: Продажа на экспорт": wrap_row_data(worksheet, 26, column),
            "Production": wrap_row_data(worksheet, 27, column),

            "Elastic": wrap_row_data(worksheet, 1, column)
        }

        plt.figure(figsize=(10, 6))
        for key in data.keys():
            if key != "Elastic":
                plt.plot(data["Elastic"], data[key], marker='o', label=key)

        plt.xlabel("Elastic")
        plt.ylabel("Млн.руб")
        plt.title("График общей чувствительности")
        plt.legend(loc='upper left', bbox_to_anchor=(1, 1))
        plt.grid(True)
        plt.savefig(path + "график2.png", bbox_inches="tight")

        # График общей чувствительности по CAPEX
        data = {
            "CAPEX: 1. БУРЕНИЕ ДОБЫВАЮЩИХ СКВАЖИН": wrap_row_data(worksheet, 3, column),
            "CAPEX: 2. БУРЕНИЕ НАГНЕТАТЕЛЬНЫХ СКВАЖИН": wrap_row_data(worksheet, 4, column),
            "CAPEX: 3. МЕХАНИЗАЦИЯ СКВАЖИН": wrap_row_data(worksheet, 5, column),
            "CAPEX: 4. СБОР И ТРАНСПОРТ НЕФТИ": wrap_row_data(worksheet, 6, column),
            "CAPEX: 5. ДНС с УПСВ": wrap_row_data(worksheet, 7, column),
            "CAPEX: 6. ЭЛЕКТРОСНАБЖЕНИЕ И СВЯЗЬ": wrap_row_data(worksheet, 8, column),
            "CAPEX: 7. ПРОМВОДОСНАБЖЕНИЕ": wrap_row_data(worksheet, 9, column),
            "CAPEX: 8. ПРОМЫСЛОВЫЕ ДОРОГИ": wrap_row_data(worksheet, 10, column),
            "CAPEX: 9.ППД": wrap_row_data(worksheet, 11, column),
            "CAPEX: 10. ЗАТРАТЫ НА ЭКОЛОГИЮ": wrap_row_data(worksheet, 12, column),
            "CAPEX: 11. ПРОЧИЕ": wrap_row_data(worksheet, 13, column),

            "Elastic": wrap_row_data(worksheet, 1, column)
        }

        plt.figure(figsize=(10, 6))
        for key in data.keys():
            if key != "Elastic":
                plt.plot(data["Elastic"], data[key], marker='o', label=key)

        plt.xlabel("Elastic")
        plt.ylabel("Млн.руб")
        plt.title("График общей чувствительности по CAPEX")
        plt.legend(loc='upper left', bbox_to_anchor=(1, 1))
        plt.grid(True)
        plt.savefig(path + "график3.png", bbox_inches="tight")

        # График общей чувствительности по OPEX
        data = {
            "OPEX: обслуживание добывающих скважин": wrap_row_data(worksheet, 15, column),
            "OPEX: обслуживание нагнетательных скважин": wrap_row_data(worksheet, 16, column),
            "OPEX: подготовка нефти": wrap_row_data(worksheet, 17, column),
            "OPEX: сбор и транспорт нефти": wrap_row_data(worksheet, 18, column),
            "OPEX: закачка воды": wrap_row_data(worksheet, 19, column),
            "OPEX: Механизированное извлечение нефти": wrap_row_data(worksheet, 20, column),
            "OPEX: ГРП": wrap_row_data(worksheet, 21, column),
            "OPEX: изоляция пласта": wrap_row_data(worksheet, 22, column),
            "OPEX: Капитальный ремонт": wrap_row_data(worksheet, 23, column),
            "OPEX: прочие": wrap_row_data(worksheet, 24, column),

            "Elastic": [0.70, 0.80, 0.90, 1.00, 1.10, 1.20, 1.30]
        }

        plt.figure(figsize=(10, 6))
        for key in data.keys():
            if key != "Elastic":
                plt.plot(data["Elastic"], data[key], marker='o', label=key)

        plt.xlabel("Elastic")
        plt.ylabel("Млн.руб")
        plt.title("График общей чувствительности по OPEX")
        plt.legend(bbox_to_anchor=(1, 1))
        plt.grid(True)
        plt.savefig(path + "график4.png", bbox_inches="tight")

    def montecarlo_to_image(self, path='Charts/'):
        wrap_column_data = ChartsToImage.wrap_column_data
        worksheet = self.workbook['MonteСarlo']
        data = {
            "Interval": wrap_column_data(worksheet, 4),
            "Quantity": list(map(lambda x: round(float(x), 2), wrap_column_data(worksheet, 5)))
        }

        # Получение интервалов и количества из данных
        intervals = data["Interval"]
        quantity = data["Quantity"]

        # Создание гистограммы
        plt.figure(figsize=(10, 6))
        plt.bar(intervals, quantity, color='skyblue')
        plt.xlabel("Интервал")
        plt.ylabel("Количество")
        plt.title("Гистограмма данных по интервалам")
        plt.xticks(rotation=45, ha='right')  # Поворот подписей на оси x для улучшения читаемости
        plt.tight_layout()  # Автоматический подбор параметров для красивого размещения элементов
        plt.savefig(path + "график5.png", bbox_inches="tight")

    def charts_to_image(self):
        ChartsToImage.npv_to_image(self)
        ChartsToImage.elastic_to_image(self)
        ChartsToImage.montecarlo_to_image(self)

    def __del__(self):
        print(f"Графики были преобразованы в изображение за {round(time.time() - self.time, 2)} с.")


if __name__ == "__main__":
    excel_file = "Results/Overview_E3.xlsx"
    wb = load_workbook(excel_file)
    ChartsToImage(wb).charts_to_image()
